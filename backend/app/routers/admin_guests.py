import secrets
import math
from io import BytesIO
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_
from sqlalchemy.orm import selectinload
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel

from app.database import get_db
from app.models import (
    Wedding, Guest, TravelInfo, HotelInfo,
    GuestFoodPreference, GuestDressPreference, GuestActivity, RSVPStatus
)
from app.models.dress_code import DressCode
from app.models.activity import Activity
from app.schemas import (
    GuestCreate, GuestResponse, GuestListResponse, SuccessResponse
)
from app.utils.auth import get_current_wedding
from app.config import settings

router = APIRouter(prefix="/api/admin/guests", tags=["Admin Guests"])


class GuestWithFlags(BaseModel):
    id: str
    wedding_id: str
    unique_token: str
    full_name: str
    first_name: str = ""
    last_name: str = ""
    email: Optional[str] = None
    phone: Optional[str] = None
    country_of_origin: Optional[str] = None
    country: Optional[str] = None
    rsvp_status: str
    number_of_attendees: int
    number_of_guests: int = 0
    special_requests: Optional[str] = None
    travel_info_submitted: bool
    hotel_info_submitted: bool
    has_travel_info: bool = False
    has_hotel_info: bool = False
    guest_link: str

    class Config:
        from_attributes = True


class BulkUploadResponse(BaseModel):
    created_count: int
    created: int = 0
    skipped_count: int
    errors: List[str] = []
    guests: List[GuestWithFlags]


class GuestUpdate(BaseModel):
    full_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    country_of_origin: Optional[str] = None
    rsvp_status: Optional[str] = None
    number_of_attendees: Optional[int] = None
    special_requests: Optional[str] = None


def generate_guest_link(unique_token: str) -> str:
    return f"{settings.FRONTEND_URL}/guest/{unique_token}"


@router.post("/upload-excel", response_model=BulkUploadResponse)
async def bulk_upload_guests(
    file: UploadFile = File(...),
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """Bulk upload guests from Excel."""
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be .xlsx, .xls, or .csv"
        )

    # Read file
    content = await file.read()

    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        else:
            df = pd.read_excel(BytesIO(content))
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error reading file: {str(e)}"
        )

    # Normalize column names
    df.columns = df.columns.str.lower().str.strip().str.replace(' ', '_')

    if 'full_name' not in df.columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must have 'full_name' column"
        )

    # Get existing guests for duplicate check
    existing_result = await db.execute(
        select(Guest.full_name).where(Guest.wedding_id == wedding.id)
    )
    existing_names = {row.lower() for row in existing_result.scalars().all()}

    created_guests = []
    skipped_count = 0

    for _, row in df.iterrows():
        full_name = str(row.get('full_name', '')).strip()
        if not full_name or full_name.lower() == 'nan':
            skipped_count += 1
            continue

        # Check for duplicate
        if full_name.lower() in existing_names:
            skipped_count += 1
            continue

        email = row.get('email') if pd.notna(row.get('email')) else None

        guest = Guest(
            wedding_id=wedding.id,
            unique_token=secrets.token_urlsafe(32),
            full_name=full_name,
            email=str(email) if email else None,
            rsvp_status=RSVPStatus.pending,
            number_of_attendees=1
        )

        db.add(guest)
        existing_names.add(full_name.lower())
        created_guests.append(guest)

    await db.flush()

    # Refresh all guests
    for guest in created_guests:
        await db.refresh(guest)

    guest_responses = []
    for g in created_guests:
        name_parts = g.full_name.split(' ', 1)
        guest_responses.append(GuestWithFlags(
            id=str(g.id),
            wedding_id=str(g.wedding_id),
            unique_token=g.unique_token,
            full_name=g.full_name,
            first_name=name_parts[0],
            last_name=name_parts[1] if len(name_parts) > 1 else '',
            email=g.email,
            phone=g.phone,
            country_of_origin=g.country_of_origin,
            country=g.country_of_origin,
            rsvp_status=g.rsvp_status.value,
            number_of_attendees=g.number_of_attendees,
            number_of_guests=g.number_of_attendees,
            special_requests=g.special_requests,
            travel_info_submitted=False,
            hotel_info_submitted=False,
            has_travel_info=False,
            has_hotel_info=False,
            guest_link=generate_guest_link(g.unique_token)
        ))

    created_count = len(created_guests)
    return BulkUploadResponse(
        created_count=created_count,
        created=created_count,
        skipped_count=skipped_count,
        errors=[],
        guests=guest_responses
    )


@router.get("/", response_model=GuestListResponse)
async def list_guests(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    rsvp_status: Optional[str] = None,
    search: Optional[str] = None,
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """List all guests with pagination and filters."""
    query = select(Guest).where(Guest.wedding_id == wedding.id)
    count_query = select(func.count(Guest.id)).where(Guest.wedding_id == wedding.id)

    # Apply filters
    if rsvp_status:
        query = query.where(Guest.rsvp_status == RSVPStatus(rsvp_status))
        count_query = count_query.where(Guest.rsvp_status == RSVPStatus(rsvp_status))

    if search:
        search_filter = or_(
            Guest.full_name.ilike(f"%{search}%"),
            Guest.email.ilike(f"%{search}%"),
            Guest.phone.ilike(f"%{search}%")
        )
        query = query.where(search_filter)
        count_query = count_query.where(search_filter)

    # Get total count
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    # Apply pagination
    offset = (page - 1) * page_size
    query = query.offset(offset).limit(page_size).order_by(Guest.created_at.desc())

    result = await db.execute(query)
    guests = result.scalars().all()

    # Check for travel and hotel info
    guest_ids = [g.id for g in guests]

    travel_result = await db.execute(
        select(TravelInfo.guest_id).where(TravelInfo.guest_id.in_(guest_ids))
    )
    travel_submitted = {row for row in travel_result.scalars().all()}

    hotel_result = await db.execute(
        select(HotelInfo.guest_id).where(HotelInfo.guest_id.in_(guest_ids))
    )
    hotel_submitted = {row for row in hotel_result.scalars().all()}

    guest_responses = [
        GuestResponse(
            id=g.id,
            wedding_id=g.wedding_id,
            unique_token=g.unique_token,
            full_name=g.full_name,
            email=g.email,
            phone=g.phone,
            country_of_origin=g.country_of_origin,
            rsvp_status=g.rsvp_status,
            number_of_attendees=g.number_of_attendees,
            special_requests=g.special_requests,
            last_accessed_at=g.last_accessed_at,
            created_at=g.created_at,
            updated_at=g.updated_at
        )
        for g in guests
    ]

    total_pages = math.ceil(total / page_size)

    return GuestListResponse(
        items=guest_responses,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
        has_next=page < total_pages,
        has_previous=page > 1
    )


@router.post("/", response_model=GuestResponse, status_code=status.HTTP_201_CREATED)
async def create_guest(
    data: GuestCreate,
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """Add single guest manually."""
    guest = Guest(
        wedding_id=wedding.id,
        unique_token=secrets.token_urlsafe(32),
        full_name=data.full_name,
        email=data.email,
        phone=data.phone,
        country_of_origin=data.country_of_origin,
        rsvp_status=RSVPStatus.pending,
        number_of_attendees=1
    )

    db.add(guest)
    await db.flush()
    await db.refresh(guest)

    return GuestResponse.model_validate(guest)


@router.get("/export")
async def export_guests(
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """Export all guest data to Excel."""
    result = await db.execute(
        select(Guest)
        .options(
            selectinload(Guest.travel_info),
            selectinload(Guest.hotel_info),
            selectinload(Guest.food_preference),
            selectinload(Guest.dress_preferences),
            selectinload(Guest.activities)
        )
        .where(Guest.wedding_id == wedding.id)
        .order_by(Guest.full_name)
    )
    guests = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Guests"

    # Headers
    headers = [
        "Full Name", "Email", "Phone", "Country", "RSVP Status",
        "Attendees", "Special Requests", "Guest Link",
        "Travel Info", "Hotel Info", "Food Preferences"
    ]

    header_fill = PatternFill(start_color="C9A961", end_color="C9A961", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row, guest in enumerate(guests, 2):
        ws.cell(row=row, column=1, value=guest.full_name)
        ws.cell(row=row, column=2, value=guest.email or "")
        ws.cell(row=row, column=3, value=guest.phone or "")
        ws.cell(row=row, column=4, value=guest.country_of_origin or "")
        ws.cell(row=row, column=5, value=guest.rsvp_status.value)
        ws.cell(row=row, column=6, value=guest.number_of_attendees)
        ws.cell(row=row, column=7, value=guest.special_requests or "")
        ws.cell(row=row, column=8, value=generate_guest_link(guest.unique_token))
        ws.cell(row=row, column=9, value="Yes" if guest.travel_info else "No")
        ws.cell(row=row, column=10, value="Yes" if guest.hotel_info else "No")
        ws.cell(row=row, column=11, value="Yes" if guest.food_preference else "No")

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=guests_export.xlsx"}
    )


@router.get("/{guest_id}")
async def get_guest(
    guest_id: str,
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """Get single guest details with all submitted data."""
    result = await db.execute(
        select(Guest)
        .options(
            selectinload(Guest.travel_info),
            selectinload(Guest.hotel_info).selectinload(HotelInfo.suggested_hotel),
            selectinload(Guest.food_preference),
            selectinload(Guest.dress_preferences).selectinload(GuestDressPreference.dress_code),
            selectinload(Guest.activities).selectinload(GuestActivity.activity)
        )
        .where(Guest.id == guest_id, Guest.wedding_id == wedding.id)
    )
    guest = result.scalar_one_or_none()

    if not guest:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Guest not found"
        )

    name_parts = guest.full_name.split(' ', 1)

    # Serialize travel info
    travel_info = None
    if guest.travel_info:
        ti = guest.travel_info
        travel_info = {
            "arrival_date": ti.arrival_date.isoformat() if ti.arrival_date else None,
            "arrival_time": ti.arrival_time,
            "arrival_flight_number": ti.arrival_flight_number,
            "arrival_airport": ti.arrival_airport,
            "departure_date": ti.departure_date.isoformat() if ti.departure_date else None,
            "departure_time": ti.departure_time,
            "departure_flight_number": ti.departure_flight_number,
            "needs_pickup": ti.needs_pickup,
            "needs_dropoff": ti.needs_dropoff,
            "special_requirements": ti.special_requirements,
            "special_requests": ti.special_requirements,
            "updated_at": ti.updated_at.isoformat() if ti.updated_at else None,
        }

    # Serialize hotel info
    hotel_info = None
    if guest.hotel_info:
        hi = guest.hotel_info
        hotel_name = hi.custom_hotel_name
        if hi.suggested_hotel:
            hotel_name = hi.suggested_hotel.hotel_name
        hotel_info = {
            "hotel_name": hotel_name or "Self-arranged",
            "custom_hotel_name": hi.custom_hotel_name,
            "custom_hotel_address": hi.custom_hotel_address,
            "check_in_date": hi.check_in_date.isoformat() if hi.check_in_date else None,
            "check_out_date": hi.check_out_date.isoformat() if hi.check_out_date else None,
            "room_type": hi.room_type,
            "number_of_rooms": hi.number_of_rooms,
            "special_requests": hi.special_requests,
            "booking_confirmation": hi.booking_confirmation,
            "is_booked": bool(hi.booking_confirmation),
        }

    # Serialize food preference
    food_preferences = None
    if guest.food_preference:
        fp = guest.food_preference
        food_preferences = {
            "dietary_restrictions": fp.dietary_restrictions,
            "allergies": fp.allergies,
            "cuisine_preferences": fp.cuisine_preferences,
            "special_requests": fp.special_requests,
            "meal_size_preference": fp.meal_size_preference.value if fp.meal_size_preference else None,
        }

    # Serialize dress preferences
    dress_preferences = []
    for dp in (guest.dress_preferences or []):
        event_name = dp.dress_code.event_name if dp.dress_code else "Unknown Event"
        dress_preferences.append({
            "event_name": event_name,
            "dress_code_id": str(dp.dress_code_id),
            "planned_outfit_description": dp.planned_outfit_description,
            "outfit_choice": dp.planned_outfit_description,
            "color_choice": dp.color_choice,
            "color_code": dp.color_choice,
            "color_name": dp.color_choice,
            "needs_shopping_assistance": dp.needs_shopping_assistance,
            "notes": dp.notes,
        })

    # Serialize activities
    activities = []
    for ga in (guest.activities or []):
        act = ga.activity
        activities.append({
            "activity_name": act.activity_name if act else "Activity",
            "name": act.activity_name if act else "Activity",
            "date": act.date_time.strftime("%Y-%m-%d") if act and act.date_time else None,
            "time": act.date_time.strftime("%I:%M %p") if act and act.date_time else None,
            "start_time": act.date_time.isoformat() if act and act.date_time else None,
            "location": act.location if act else None,
            "number_of_participants": ga.number_of_participants,
            "notes": ga.notes,
            "registered_at": ga.registered_at.isoformat() if ga.registered_at else None,
        })

    return {
        "id": str(guest.id),
        "wedding_id": str(guest.wedding_id),
        "unique_token": guest.unique_token,
        "full_name": guest.full_name,
        "first_name": name_parts[0],
        "last_name": name_parts[1] if len(name_parts) > 1 else "",
        "email": guest.email,
        "phone": guest.phone,
        "country_of_origin": guest.country_of_origin,
        "country": guest.country_of_origin,
        "rsvp_status": guest.rsvp_status.value,
        "number_of_attendees": guest.number_of_attendees,
        "number_of_guests": guest.number_of_attendees,
        "special_requests": guest.special_requests,
        "song_requests": guest.song_requests,
        "notes_to_couple": guest.notes_to_couple,
        "rsvp_submitted_at": guest.rsvp_submitted_at.isoformat() if guest.rsvp_submitted_at else None,
        "created_at": guest.created_at.isoformat() if guest.created_at else None,
        "updated_at": guest.updated_at.isoformat() if guest.updated_at else None,
        "last_accessed_at": guest.last_accessed_at.isoformat() if guest.last_accessed_at else None,
        "travel_info_submitted": guest.travel_info is not None,
        "hotel_info_submitted": guest.hotel_info is not None,
        "has_travel_info": guest.travel_info is not None,
        "has_hotel_info": guest.hotel_info is not None,
        "guest_link": generate_guest_link(guest.unique_token),
        "travel_info": travel_info,
        "hotel_info": hotel_info,
        "food_preferences": food_preferences,
        "dress_preferences": dress_preferences,
        "activities": activities,
    }


@router.put("/{guest_id}", response_model=GuestResponse)
async def update_guest(
    guest_id: str,
    data: GuestUpdate,
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """Update guest information."""
    result = await db.execute(
        select(Guest).where(Guest.id == guest_id, Guest.wedding_id == wedding.id)
    )
    guest = result.scalar_one_or_none()

    if not guest:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Guest not found"
        )

    update_data = data.model_dump(exclude_unset=True)
    if 'rsvp_status' in update_data and update_data['rsvp_status']:
        update_data['rsvp_status'] = RSVPStatus(update_data['rsvp_status'])

    for key, value in update_data.items():
        setattr(guest, key, value)

    await db.flush()
    await db.refresh(guest)

    return GuestResponse.model_validate(guest)


@router.delete("/{guest_id}", response_model=SuccessResponse)
async def delete_guest(
    guest_id: str,
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """Delete a guest."""
    result = await db.execute(
        select(Guest).where(Guest.id == guest_id, Guest.wedding_id == wedding.id)
    )
    guest = result.scalar_one_or_none()

    if not guest:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Guest not found"
        )

    await db.delete(guest)
    await db.flush()

    return SuccessResponse(message="Guest deleted successfully")


@router.post("/{guest_id}/regenerate-link", response_model=GuestWithFlags)
async def regenerate_guest_link(
    guest_id: str,
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """Regenerate unique token/link for a guest."""
    result = await db.execute(
        select(Guest).where(Guest.id == guest_id, Guest.wedding_id == wedding.id)
    )
    guest = result.scalar_one_or_none()

    if not guest:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Guest not found"
        )

    guest.unique_token = secrets.token_urlsafe(32)
    await db.flush()
    await db.refresh(guest)

    name_parts = guest.full_name.split(' ', 1)
    return GuestWithFlags(
        id=str(guest.id),
        wedding_id=str(guest.wedding_id),
        unique_token=guest.unique_token,
        full_name=guest.full_name,
        first_name=name_parts[0],
        last_name=name_parts[1] if len(name_parts) > 1 else '',
        email=guest.email,
        phone=guest.phone,
        country_of_origin=guest.country_of_origin,
        country=guest.country_of_origin,
        rsvp_status=guest.rsvp_status.value,
        number_of_attendees=guest.number_of_attendees,
        number_of_guests=guest.number_of_attendees,
        special_requests=guest.special_requests,
        travel_info_submitted=False,
        hotel_info_submitted=False,
        has_travel_info=False,
        has_hotel_info=False,
        guest_link=generate_guest_link(guest.unique_token)
    )
