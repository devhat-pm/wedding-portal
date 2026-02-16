from io import BytesIO
from typing import List
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app.models import Wedding, FoodMenu, Guest, GuestFoodPreference
from app.schemas import (
    FoodMenuCreate,
    FoodMenuUpdate,
    FoodMenuResponse,
    GuestFoodPreferenceResponse,
    SuccessResponse
)
from app.utils.auth import get_current_wedding

router = APIRouter(prefix="/api/admin/food-menu", tags=["Admin Food Menu"])


@router.post("/", response_model=FoodMenuResponse, status_code=status.HTTP_201_CREATED)
async def create_food_menu(
    data: FoodMenuCreate,
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """Create food menu for event."""
    food_menu = FoodMenu(
        wedding_id=wedding.id,
        event_name=data.event_name,
        menu_items=data.menu_items,
        dietary_options_available=data.dietary_options_available,
        notes=data.notes
    )

    db.add(food_menu)
    await db.flush()
    await db.refresh(food_menu)

    return FoodMenuResponse.model_validate(food_menu)


@router.get("/", response_model=List[FoodMenuResponse])
async def list_food_menus(
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """List all menus."""
    result = await db.execute(
        select(FoodMenu).where(FoodMenu.wedding_id == wedding.id)
    )
    menus = result.scalars().all()

    return [FoodMenuResponse.model_validate(m) for m in menus]


@router.put("/{menu_id}", response_model=FoodMenuResponse)
async def update_food_menu(
    menu_id: UUID,
    data: FoodMenuUpdate,
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """Update menu."""
    result = await db.execute(
        select(FoodMenu).where(
            FoodMenu.id == menu_id,
            FoodMenu.wedding_id == wedding.id
        )
    )
    menu = result.scalar_one_or_none()

    if not menu:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Menu not found"
        )

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(menu, field, value)

    await db.flush()
    await db.refresh(menu)

    return FoodMenuResponse.model_validate(menu)


@router.delete("/{menu_id}", response_model=SuccessResponse)
async def delete_food_menu(
    menu_id: UUID,
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """Remove menu."""
    result = await db.execute(
        select(FoodMenu).where(
            FoodMenu.id == menu_id,
            FoodMenu.wedding_id == wedding.id
        )
    )
    menu = result.scalar_one_or_none()

    if not menu:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Menu not found"
        )

    await db.delete(menu)
    await db.flush()

    return SuccessResponse(message="Menu deleted successfully")


@router.get("/guest-preferences", response_model=List[GuestFoodPreferenceResponse])
async def list_guest_food_preferences(
    export: bool = Query(default=False),
    wedding: Wedding = Depends(get_current_wedding),
    db: AsyncSession = Depends(get_db)
):
    """List all guest food preferences with export option."""
    result = await db.execute(
        select(GuestFoodPreference)
        .join(Guest)
        .where(Guest.wedding_id == wedding.id)
        .options(selectinload(GuestFoodPreference.guest))
    )
    preferences = result.scalars().all()

    if export:
        wb = Workbook()
        ws = wb.active
        ws.title = "Food Preferences"

        headers = [
            "Guest Name", "Dietary Restrictions", "Allergies",
            "Cuisine Preferences", "Special Requests", "Meal Size"
        ]

        header_fill = PatternFill(start_color="C9A961", end_color="C9A961", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row, pref in enumerate(preferences, 2):
            ws.cell(row=row, column=1, value=pref.guest.full_name if pref.guest else "")
            ws.cell(row=row, column=2, value=", ".join(pref.dietary_restrictions) if pref.dietary_restrictions else "")
            ws.cell(row=row, column=3, value=pref.allergies or "")
            ws.cell(row=row, column=4, value=pref.cuisine_preferences or "")
            ws.cell(row=row, column=5, value=pref.special_requests or "")
            ws.cell(row=row, column=6, value=pref.meal_size_preference.value if pref.meal_size_preference else "")

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=food_preferences.xlsx"}
        )

    return [GuestFoodPreferenceResponse.model_validate(p) for p in preferences]
