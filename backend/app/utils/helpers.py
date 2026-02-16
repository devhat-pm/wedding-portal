import uuid
import re
from typing import Optional, BinaryIO
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def generate_invitation_code(length: int = 8) -> str:
    """Generate a unique invitation code."""
    return str(uuid.uuid4())[:length].upper()


def format_phone_number(phone: Optional[str]) -> Optional[str]:
    """Format phone number to a consistent format."""
    if not phone:
        return None

    # Remove all non-numeric characters except +
    cleaned = re.sub(r'[^\d+]', '', phone)
    return cleaned if cleaned else None


def export_guests_to_excel(guests: list[dict]) -> BytesIO:
    """Export guests to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Guests"

    # Define headers
    headers = [
        "First Name", "Last Name", "First Name (Arabic)", "Last Name (Arabic)",
        "Email", "Phone", "Side", "Relation", "RSVP Status",
        "Plus One", "Plus One Name", "Children Count",
        "Table #", "Seat #", "VIP", "Dietary Restrictions",
        "Special Requests", "Notes"
    ]

    # Style for header
    header_fill = PatternFill(start_color="C9A961", end_color="C9A961", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row, guest in enumerate(guests, 2):
        ws.cell(row=row, column=1, value=guest.get("first_name", ""))
        ws.cell(row=row, column=2, value=guest.get("last_name", ""))
        ws.cell(row=row, column=3, value=guest.get("first_name_arabic", ""))
        ws.cell(row=row, column=4, value=guest.get("last_name_arabic", ""))
        ws.cell(row=row, column=5, value=guest.get("email", ""))
        ws.cell(row=row, column=6, value=guest.get("phone", ""))
        ws.cell(row=row, column=7, value=guest.get("side", ""))
        ws.cell(row=row, column=8, value=guest.get("relation", ""))
        ws.cell(row=row, column=9, value=guest.get("rsvp_status", ""))
        ws.cell(row=row, column=10, value="Yes" if guest.get("plus_one_allowed") else "No")
        ws.cell(row=row, column=11, value=guest.get("plus_one_name", ""))
        ws.cell(row=row, column=12, value=guest.get("children_count", 0))
        ws.cell(row=row, column=13, value=guest.get("table_number", ""))
        ws.cell(row=row, column=14, value=guest.get("seat_number", ""))
        ws.cell(row=row, column=15, value="Yes" if guest.get("is_vip") else "No")
        ws.cell(row=row, column=16, value=guest.get("dietary_restrictions", ""))
        ws.cell(row=row, column=17, value=guest.get("special_requests", ""))
        ws.cell(row=row, column=18, value=guest.get("notes", ""))

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_guests_from_excel(file: BinaryIO) -> list[dict]:
    """Import guests from Excel file."""
    df = pd.read_excel(file)

    # Normalize column names
    df.columns = df.columns.str.lower().str.replace(' ', '_').str.replace('(', '').str.replace(')', '')

    # Column mapping
    column_mapping = {
        'first_name': 'first_name',
        'last_name': 'last_name',
        'first_name_arabic': 'first_name_arabic',
        'last_name_arabic': 'last_name_arabic',
        'email': 'email',
        'phone': 'phone',
        'side': 'side',
        'relation': 'relation',
        'vip': 'is_vip',
        'plus_one': 'plus_one_allowed',
        'notes': 'notes'
    }

    guests = []
    for _, row in df.iterrows():
        guest = {}
        for excel_col, schema_col in column_mapping.items():
            if excel_col in df.columns:
                value = row[excel_col]
                if pd.isna(value):
                    value = None
                elif excel_col in ['vip', 'plus_one']:
                    value = str(value).lower() in ['yes', 'true', '1', 'y']
                guest[schema_col] = value

        # Only add if we have at least first and last name
        if guest.get('first_name') and guest.get('last_name'):
            guests.append(guest)

    return guests


def generate_guest_template() -> BytesIO:
    """Generate an Excel template for importing guests."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Guest Template"

    # Define headers
    headers = [
        "First Name*", "Last Name*", "First Name (Arabic)", "Last Name (Arabic)",
        "Email", "Phone", "Side (bride/groom/both)", "Relation (family/friend/colleague/neighbor/other)",
        "VIP (Yes/No)", "Plus One (Yes/No)", "Notes"
    ]

    # Style for header
    header_fill = PatternFill(start_color="C9A961", end_color="C9A961", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = 20

    # Add sample data
    sample_data = [
        ["Ahmed", "Al-Rashid", "أحمد", "الراشد", "ahmed@example.com", "+966501234567", "groom", "family", "Yes", "Yes", "VIP Guest"],
        ["Fatima", "Hassan", "فاطمة", "حسن", "fatima@example.com", "+966509876543", "bride", "friend", "No", "No", ""]
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
